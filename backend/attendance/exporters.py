import csv
from io import BytesIO, StringIO

from django.utils import timezone

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from openpyxl import Workbook
from openpyxl.styles import Font

HEADERS = ["Employee", "Email", "Days present", "Hours worked", "Leave days"]


def _row_values(row):
    return [row["name"], row["email"], row["days_present"], round(row["hours_worked"], 2), row["leave_days"]]


def render_pdf(rows, date_from, date_to):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    elements = [
        Paragraph(f"Attendance report &middot; {date_from} to {date_to}", styles["Title"]),
        Spacer(1, 12),
    ]

    data = [HEADERS] + [_row_values(r) for r in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16233A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E1E3DC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F6F2")]),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)

    return buffer.getvalue()


def render_excel(rows, date_from, date_to):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance report"

    ws.append([f"Attendance report", f"{date_from} to {date_to}"])
    ws.append([])
    ws.append(HEADERS)
    for cell in ws[3]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(_row_values(row))

    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(40, length + 2)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()




TIMESHEET_HEADERS = ["Employee", "Date", "Scheduled shift", "Check-in", "Check-out", "Hours worked", "Overtime", "Shortfall", "Status"]


def _timesheet_row_values(row):
    return [
        row["name"],
        row["date"].isoformat(),
        row["scheduled_shift"] or "—",
        timezone.localtime(row["check_in"]).strftime("%H:%M") if row["check_in"] else "—",
        timezone.localtime(row["check_out"]).strftime("%H:%M") if row["check_out"] else "—",
        round(row["worked_hours"], 2),
        round(row["overtime_hours"], 2),
        round(row["shortfall_hours"], 2),
        row["status"],
    ]


def render_timesheet_csv(rows):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(TIMESHEET_HEADERS)
    for row in rows:
        writer.writerow(_timesheet_row_values(row))
    return buffer.getvalue().encode("utf-8-sig")


def render_timesheet_pdf(rows, date_from, date_to):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    elements = [
        Paragraph(f"Timesheet &middot; {date_from} to {date_to}", styles["Title"]),
        Spacer(1, 12),
    ]

    data = [TIMESHEET_HEADERS] + [_timesheet_row_values(r) for r in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16233A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E1E3DC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F6F2")]),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)

    return buffer.getvalue()


def render_timesheet_excel(rows, date_from, date_to):
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    ws.append(["Timesheet", f"{date_from} to {date_to}"])
    ws.append([])
    ws.append(TIMESHEET_HEADERS)
    for cell in ws[3]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(_timesheet_row_values(row))

    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(40, length + 2)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
